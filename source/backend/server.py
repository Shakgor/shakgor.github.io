from dotenv import load_dotenv
from pathlib import Path

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

import os
import io
import json
import uuid
import zipfile
import logging
import bcrypt
import jwt
from datetime import datetime, timezone, timedelta
from typing import List, Optional, Any, Dict

from fastapi import FastAPI, APIRouter, HTTPException, Request, Response, Depends, UploadFile, File
from fastapi.responses import StreamingResponse, JSONResponse
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, EmailStr, Field

import openpyxl
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak

# ---------- Mongo ----------
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

JWT_ALGORITHM = "HS256"

def now_utc() -> datetime:
    return datetime.now(timezone.utc)

def jwt_secret() -> str:
    return os.environ["JWT_SECRET"]

def hash_password(p: str) -> str:
    return bcrypt.hashpw(p.encode(), bcrypt.gensalt()).decode()

def verify_password(p: str, h: str) -> bool:
    try:
        return bcrypt.checkpw(p.encode(), h.encode())
    except Exception:
        return False

def create_token(user_id: str, role: str, ttl_minutes: int = 60 * 24 * 7) -> str:
    payload = {"sub": user_id, "role": role,
               "exp": now_utc() + timedelta(minutes=ttl_minutes), "type": "access"}
    return jwt.encode(payload, jwt_secret(), algorithm=JWT_ALGORITHM)

def serialize(doc: Dict[str, Any]) -> Dict[str, Any]:
    if not doc:
        return doc
    doc.pop("_id", None)
    doc.pop("password_hash", None)
    return doc

# ---------- Models ----------
class RegisterOwner(BaseModel):
    name: str
    email: EmailStr
    phone: str
    password: str
    gym_name: str

class RegisterMember(BaseModel):
    name: str
    email: EmailStr
    phone: str
    password: str
    gym_id: str

class LoginIn(BaseModel):
    identifier: str
    password: str

class ChangeCredentialsIn(BaseModel):
    new_username: Optional[str] = None
    new_password: Optional[str] = None

class GymBrandingIn(BaseModel):
    name: Optional[str] = None
    logo_url: Optional[str] = None
    primary_color: Optional[str] = None
    secondary_color: Optional[str] = None
    background_color: Optional[str] = None

class EquipmentIn(BaseModel):
    items: List[str]

class QuestionOption(BaseModel):
    label: str
    tags: List[str] = []  # e.g. ["weight_loss", "vegetarian"]
    calorie_modifier: int = 0  # added to base calorie target
    level: Optional[str] = None  # "beginner"|"intermediate"|"advanced"

class QuestionIn(BaseModel):
    text: str
    type: str = "single"  # single | number
    options: List[QuestionOption] = []
    order: int = 0
    affects: str = "general"  # "diet"|"routine"|"general"

class RoutineExercise(BaseModel):
    name: str
    sets: int = 3
    reps: str = "10-12"
    equipment: Optional[str] = None
    notes: Optional[str] = None

class RoutineIn(BaseModel):
    name: str
    level: str = "beginner"
    goal_tags: List[str] = []
    required_equipment: List[str] = []
    exercises: List[RoutineExercise] = []

class PaymentIn(BaseModel):
    member_id: str
    plan_type: str  # "monthly"|"15days"|"visit"
    amount: float
    paid_at: Optional[str] = None  # iso

class QuestionnaireSubmit(BaseModel):
    answers: List[Dict[str, Any]]  # [{question_id, option_index OR value}]
    base_calories: Optional[int] = 2000

# ---------- App ----------
app = FastAPI()
api = APIRouter(prefix="/api")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)

logger = logging.getLogger("gymbros")
logging.basicConfig(level=logging.INFO)

# ---------- Auth helpers ----------
async def get_current_user(request: Request) -> Dict[str, Any]:
    auth = request.headers.get("Authorization", "")
    token = None
    if auth.startswith("Bearer "):
        token = auth[7:]
    if not token:
        token = request.cookies.get("access_token")
    if not token:
        raise HTTPException(401, "Not authenticated")
    try:
        payload = jwt.decode(token, jwt_secret(), algorithms=[JWT_ALGORITHM])
    except jwt.ExpiredSignatureError:
        raise HTTPException(401, "Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(401, "Invalid token")
    user = await db.users.find_one({"id": payload["sub"]})
    if not user:
        raise HTTPException(401, "User not found")
    return serialize(user)

def require_role(*roles):
    async def _dep(user=Depends(get_current_user)):
        if user.get("role") not in roles:
            raise HTTPException(403, "Forbidden")
        return user
    return _dep

# ---------- Startup ----------
@app.on_event("startup")
async def startup():
    await db.users.create_index("email")
    await db.users.create_index("username")
    await db.users.create_index("id", unique=True)
    await db.gyms.create_index("id", unique=True)
    await db.gyms.create_index("slug", unique=True)
    await db.questions.create_index("order")
    await db.foods.create_index("name")
    await db.routines.create_index("level")
    await db.payments.create_index("member_id")
    await db.member_plans.create_index("member_id", unique=True)
    # Seed admin
    admin_username = os.environ.get("ADMIN_USERNAME", "admin")
    admin_password = os.environ.get("ADMIN_PASSWORD", "gymbros2026")
    existing = await db.users.find_one({"role": "admin"})
    if not existing:
        await db.users.insert_one({
            "id": str(uuid.uuid4()),
            "username": admin_username,
            "name": "Administrator",
            "role": "admin",
            "password_hash": hash_password(admin_password),
            "created_at": now_utc().isoformat(),
        })
        logger.info("Seeded admin user")
    # Seed default questions if empty
    if await db.questions.count_documents({}) == 0:
        defaults = [
            {"text": "¿Cuál es tu objetivo principal?", "type": "single",
             "affects": "general", "order": 1,
             "options": [
                 {"label": "Bajar de peso", "tags": ["weight_loss"], "calorie_modifier": -300, "level": "beginner"},
                 {"label": "Ganar masa muscular", "tags": ["muscle_gain"], "calorie_modifier": 300, "level": "intermediate"},
                 {"label": "Mantenerme en forma", "tags": ["maintenance"], "calorie_modifier": 0, "level": "beginner"},
             ]},
            {"text": "¿Cuál es tu nivel de experiencia?", "type": "single",
             "affects": "routine", "order": 2,
             "options": [
                 {"label": "Principiante", "tags": [], "calorie_modifier": 0, "level": "beginner"},
                 {"label": "Intermedio", "tags": [], "calorie_modifier": 0, "level": "intermediate"},
                 {"label": "Avanzado", "tags": [], "calorie_modifier": 0, "level": "advanced"},
             ]},
            {"text": "¿Tienes alguna restricción alimentaria?", "type": "single",
             "affects": "diet", "order": 3,
             "options": [
                 {"label": "Ninguna", "tags": [], "calorie_modifier": 0},
                 {"label": "Vegetariano", "tags": ["vegetarian"], "calorie_modifier": 0},
                 {"label": "Vegano", "tags": ["vegan"], "calorie_modifier": 0},
             ]},
        ]
        for d in defaults:
            d["id"] = str(uuid.uuid4())
            await db.questions.insert_one(d)

@app.on_event("shutdown")
async def shutdown():
    client.close()

# ---------- AUTH ----------
@api.post("/auth/register-owner")
async def register_owner(body: RegisterOwner, response: Response):
    email = body.email.lower()
    if await db.users.find_one({"email": email}):
        raise HTTPException(400, "Email ya registrado")
    user_id = str(uuid.uuid4())
    gym_id = str(uuid.uuid4())
    slug = body.gym_name.lower().replace(" ", "-")[:40] + "-" + gym_id[:6]
    await db.gyms.insert_one({
        "id": gym_id,
        "slug": slug,
        "owner_id": user_id,
        "name": body.gym_name,
        "logo_url": "",
        "primary_color": "#FF3B30",
        "secondary_color": "#007AFF",
        "background_color": "#0A0A0A",
        "equipment": [],
        "subscription_status": "pending",  # pending | active | suspended
        "subscription_expires_at": None,
        "created_at": now_utc().isoformat(),
    })
    await db.users.insert_one({
        "id": user_id,
        "email": email,
        "phone": body.phone,
        "name": body.name,
        "role": "owner",
        "gym_id": gym_id,
        "password_hash": hash_password(body.password),
        "created_at": now_utc().isoformat(),
    })
    token = create_token(user_id, "owner")
    response.set_cookie("access_token", token, httponly=True, samesite="lax", max_age=604800, path="/")
    return {"token": token, "user": {"id": user_id, "email": email, "name": body.name, "role": "owner", "gym_id": gym_id}, "gym_slug": slug}

@api.post("/auth/register-member")
async def register_member(body: RegisterMember, response: Response):
    email = body.email.lower()
    if await db.users.find_one({"email": email}):
        raise HTTPException(400, "Email ya registrado")
    gym = await db.gyms.find_one({"id": body.gym_id})
    if not gym:
        raise HTTPException(404, "Gimnasio no encontrado")
    user_id = str(uuid.uuid4())
    await db.users.insert_one({
        "id": user_id,
        "email": email,
        "phone": body.phone,
        "name": body.name,
        "role": "member",
        "gym_id": body.gym_id,
        "password_hash": hash_password(body.password),
        "created_at": now_utc().isoformat(),
        "level": "beginner",
        "level_updated_at": now_utc().isoformat(),
    })
    token = create_token(user_id, "member")
    response.set_cookie("access_token", token, httponly=True, samesite="lax", max_age=604800, path="/")
    return {"token": token, "user": {"id": user_id, "email": email, "name": body.name, "role": "member", "gym_id": body.gym_id}}

@api.post("/auth/login")
async def login(body: LoginIn, response: Response):
    ident = body.identifier.strip().lower()
    user = await db.users.find_one({"$or": [{"email": ident}, {"username": body.identifier.strip()}]})
    if not user or not verify_password(body.password, user["password_hash"]):
        raise HTTPException(401, "Credenciales inválidas")
    token = create_token(user["id"], user["role"])
    response.set_cookie("access_token", token, httponly=True, samesite="lax", max_age=604800, path="/")
    return {"token": token, "user": serialize(dict(user))}

@api.post("/auth/logout")
async def logout(response: Response):
    response.delete_cookie("access_token", path="/")
    return {"ok": True}

@api.get("/auth/me")
async def me(user=Depends(get_current_user)):
    return user

@api.post("/auth/change-credentials")
async def change_credentials(body: ChangeCredentialsIn, user=Depends(require_role("admin"))):
    update = {}
    if body.new_username:
        update["username"] = body.new_username.strip()
    if body.new_password:
        update["password_hash"] = hash_password(body.new_password)
    if update:
        await db.users.update_one({"id": user["id"]}, {"$set": update})
    return {"ok": True}

# ---------- ADMIN ----------
@api.get("/admin/owners")
async def list_owners(user=Depends(require_role("admin"))):
    out = []
    async for o in db.users.find({"role": "owner"}, {"_id": 0, "password_hash": 0}):
        gym = await db.gyms.find_one({"id": o.get("gym_id")}, {"_id": 0})
        o["gym"] = gym
        out.append(o)
    return out

@api.post("/admin/gyms/{gym_id}/subscription")
async def update_subscription(gym_id: str, body: Dict[str, Any], user=Depends(require_role("admin"))):
    status = body.get("status", "active")  # active | suspended | pending
    days = int(body.get("days", 30))
    expires = (now_utc() + timedelta(days=days)).isoformat() if status == "active" else None
    await db.gyms.update_one({"id": gym_id}, {"$set": {
        "subscription_status": status,
        "subscription_expires_at": expires,
    }})
    return {"ok": True, "status": status, "expires_at": expires}

# Questions CRUD (admin)
@api.get("/admin/questions")
async def admin_list_questions(user=Depends(require_role("admin"))):
    return await db.questions.find({}, {"_id": 0}).sort("order", 1).to_list(500)

@api.post("/admin/questions")
async def admin_create_question(body: QuestionIn, user=Depends(require_role("admin"))):
    doc = body.model_dump()
    doc["id"] = str(uuid.uuid4())
    await db.questions.insert_one(doc)
    return serialize(doc)

@api.put("/admin/questions/{qid}")
async def admin_update_question(qid: str, body: QuestionIn, user=Depends(require_role("admin"))):
    await db.questions.update_one({"id": qid}, {"$set": body.model_dump()})
    return {"ok": True}

@api.delete("/admin/questions/{qid}")
async def admin_delete_question(qid: str, user=Depends(require_role("admin"))):
    await db.questions.delete_one({"id": qid})
    return {"ok": True}

# Foods (xlsx upload + list)
@api.post("/admin/foods/upload")
async def admin_upload_foods(file: UploadFile = File(...), user=Depends(require_role("admin"))):
    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    ws = wb.active
    headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
    required = {"name", "calories", "amount", "unit"}
    if not required.issubset(set(headers)):
        raise HTTPException(400, f"Headers requeridos: name, calories, amount, unit, [tags], [meal_type]")
    await db.foods.delete_many({})
    inserted = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = dict(zip(headers, row))
        if not rec.get("name"):
            continue
        food = {
            "id": str(uuid.uuid4()),
            "name": str(rec.get("name")),
            "calories": float(rec.get("calories") or 0),
            "amount": float(rec.get("amount") or 0),
            "unit": str(rec.get("unit") or "g"),
            "tags": [t.strip() for t in str(rec.get("tags") or "").split(",") if t.strip()],
            "meal_type": str(rec.get("meal_type") or "any"),
        }
        await db.foods.insert_one(food)
        inserted += 1
    return {"ok": True, "count": inserted}

@api.get("/admin/foods")
async def admin_list_foods(user=Depends(require_role("admin"))):
    return await db.foods.find({}, {"_id": 0}).to_list(2000)

@api.delete("/admin/foods")
async def admin_clear_foods(user=Depends(require_role("admin"))):
    await db.foods.delete_many({})
    return {"ok": True}

# Routines CRUD
@api.get("/admin/routines")
async def admin_list_routines(user=Depends(require_role("admin"))):
    return await db.routines.find({}, {"_id": 0}).to_list(500)

@api.post("/admin/routines")
async def admin_create_routine(body: RoutineIn, user=Depends(require_role("admin"))):
    doc = body.model_dump()
    doc["id"] = str(uuid.uuid4())
    await db.routines.insert_one(doc)
    return serialize(doc)

@api.delete("/admin/routines/{rid}")
async def admin_delete_routine(rid: str, user=Depends(require_role("admin"))):
    await db.routines.delete_one({"id": rid})
    return {"ok": True}

# Export members data (xlsx)
@api.get("/admin/export/members")
async def export_members(user=Depends(require_role("admin"))):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Members"
    ws.append(["Member ID", "Name", "Email", "Phone", "Gym", "Level", "Created", "Last Payment", "Status"])
    async for m in db.users.find({"role": "member"}, {"_id": 0, "password_hash": 0}):
        gym = await db.gyms.find_one({"id": m.get("gym_id")}, {"_id": 0}) or {}
        last_pay = await db.payments.find({"member_id": m["id"]}, {"_id": 0}).sort("paid_at", -1).to_list(1)
        last = last_pay[0] if last_pay else {}
        status = await _membership_status(m["id"])
        ws.append([
            m.get("id"), m.get("name"), m.get("email"), m.get("phone"),
            gym.get("name", ""), m.get("level", ""),
            m.get("created_at", ""),
            last.get("paid_at", ""),
            status.get("status", ""),
        ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=members.xlsx"})

# ZIP backup
@api.get("/admin/backup/zip")
async def backup_zip(user=Depends(require_role("admin"))):
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for col in ["users", "gyms", "questions", "foods", "routines", "payments", "member_plans"]:
            data = await db[col].find({}, {"_id": 0, "password_hash": 0}).to_list(10000)
            zf.writestr(f"data/{col}.json", json.dumps(data, indent=2, default=str))
        # include source code
        for root, _, files in os.walk("/app"):
            if any(x in root for x in ["node_modules", ".git", "__pycache__", "build", ".venv", "test_reports"]):
                continue
            for f in files:
                if f.endswith((".pyc", ".log")):
                    continue
                fp = os.path.join(root, f)
                try:
                    if os.path.getsize(fp) > 5_000_000:
                        continue
                    arc = os.path.relpath(fp, "/app")
                    zf.write(fp, f"source/{arc}")
                except Exception:
                    pass
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/zip",
                             headers={"Content-Disposition": "attachment; filename=gymbros_backup.zip"})

# ---------- OWNER ----------
@api.get("/owner/gym")
async def owner_get_gym(user=Depends(require_role("owner"))):
    gym = await db.gyms.find_one({"id": user["gym_id"]}, {"_id": 0})
    return gym

@api.put("/owner/gym/branding")
async def owner_update_branding(body: GymBrandingIn, user=Depends(require_role("owner"))):
    upd = {k: v for k, v in body.model_dump().items() if v is not None}
    await db.gyms.update_one({"id": user["gym_id"]}, {"$set": upd})
    return {"ok": True}

@api.put("/owner/gym/equipment")
async def owner_update_equipment(body: EquipmentIn, user=Depends(require_role("owner"))):
    await db.gyms.update_one({"id": user["gym_id"]}, {"$set": {"equipment": body.items}})
    return {"ok": True}

@api.get("/owner/members")
async def owner_list_members(user=Depends(require_role("owner"))):
    out = []
    async for m in db.users.find({"role": "member", "gym_id": user["gym_id"]}, {"_id": 0, "password_hash": 0}):
        status = await _membership_status(m["id"])
        m["membership"] = status
        out.append(m)
    return out

@api.post("/owner/payments")
async def owner_register_payment(body: PaymentIn, user=Depends(require_role("owner"))):
    member = await db.users.find_one({"id": body.member_id})
    if not member or member.get("gym_id") != user["gym_id"]:
        raise HTTPException(404, "Miembro no encontrado")
    days_map = {"monthly": 30, "15days": 15, "visit": 1}
    days = days_map.get(body.plan_type, 30)
    paid_at = body.paid_at or now_utc().isoformat()
    expires_at = (datetime.fromisoformat(paid_at.replace("Z", "+00:00")) + timedelta(days=days)).isoformat()
    doc = {
        "id": str(uuid.uuid4()),
        "member_id": body.member_id,
        "gym_id": user["gym_id"],
        "plan_type": body.plan_type,
        "amount": body.amount,
        "paid_at": paid_at,
        "expires_at": expires_at,
    }
    await db.payments.insert_one(doc)
    return serialize(doc)

@api.get("/owner/payments")
async def owner_list_payments(user=Depends(require_role("owner"))):
    return await db.payments.find({"gym_id": user["gym_id"]}, {"_id": 0}).sort("paid_at", -1).to_list(2000)

# ---------- PUBLIC ----------
async def _membership_status(member_id: str) -> Dict[str, Any]:
    last_pay = await db.payments.find({"member_id": member_id}, {"_id": 0}).sort("paid_at", -1).to_list(1)
    if not last_pay:
        return {"status": "no_payments", "expires_at": None}
    p = last_pay[0]
    try:
        exp = datetime.fromisoformat(p["expires_at"].replace("Z", "+00:00"))
    except Exception:
        exp = now_utc()
    if exp > now_utc():
        return {"status": "active", "expires_at": p["expires_at"], "plan_type": p["plan_type"]}
    return {"status": "expired", "expires_at": p["expires_at"], "plan_type": p["plan_type"]}

@api.get("/public/membership-check")
async def membership_check(email: str):
    user = await db.users.find_one({"email": email.lower(), "role": "member"})
    if not user:
        raise HTTPException(404, "Miembro no encontrado")
    status = await _membership_status(user["id"])
    gym = await db.gyms.find_one({"id": user.get("gym_id")}, {"_id": 0}) or {}
    return {"name": user["name"], "email": user["email"], "gym": gym.get("name", ""),
            "level": user.get("level", "beginner"), "membership": status}

@api.get("/public/gym/{slug}")
async def public_gym(slug: str):
    gym = await db.gyms.find_one({"slug": slug}, {"_id": 0})
    if not gym:
        raise HTTPException(404, "Gimnasio no encontrado")
    return gym

@api.get("/public/gyms")
async def public_gyms_list():
    return await db.gyms.find({"subscription_status": "active"}, {"_id": 0, "id": 1, "name": 1, "slug": 1}).to_list(500)

@api.get("/public/questions")
async def public_questions():
    return await db.questions.find({}, {"_id": 0}).sort("order", 1).to_list(500)

# ---------- MEMBER ----------
def _generate_diet(foods: List[Dict], target_cal: int, tags: List[str]) -> Dict:
    # Filter foods by tags (vegetarian/vegan exclusions)
    def matches(f):
        if "vegan" in tags and any(t in f.get("tags", []) for t in ["meat", "dairy", "egg"]):
            return False
        if "vegetarian" in tags and "meat" in f.get("tags", []):
            return False
        return True

    pool = [f for f in foods if matches(f)] or foods
    meals = {"breakfast": [], "lunch": [], "dinner": [], "snack": []}
    per_meal = {"breakfast": target_cal * 0.25, "lunch": target_cal * 0.35,
                "dinner": target_cal * 0.30, "snack": target_cal * 0.10}
    for meal, cal_target in per_meal.items():
        candidates = [f for f in pool if f.get("meal_type", "any") in (meal, "any")]
        accumulated = 0
        for f in candidates:
            if accumulated >= cal_target:
                break
            meals[meal].append(f)
            accumulated += f.get("calories", 0)
    return {"target_calories": target_cal, "meals": meals}

def _select_routine(routines: List[Dict], level: str, tags: List[str], available_equipment: List[str]) -> Optional[Dict]:
    avail = set([e.lower() for e in available_equipment])
    candidates = []
    for r in routines:
        if r.get("level") != level:
            continue
        req = set([e.lower() for e in r.get("required_equipment", []) or []])
        if req and not req.issubset(avail | {""}):
            continue
        score = len(set(r.get("goal_tags", [])) & set(tags))
        candidates.append((score, r))
    if not candidates:
        # Fallback: any routine matching level
        candidates = [(0, r) for r in routines if r.get("level") == level]
    if not candidates:
        return routines[0] if routines else None
    candidates.sort(key=lambda x: -x[0])
    return candidates[0][1]

@api.post("/member/questionnaire")
async def member_submit(body: QuestionnaireSubmit, user=Depends(require_role("member"))):
    questions = await db.questions.find({}, {"_id": 0}).sort("order", 1).to_list(500)
    q_map = {q["id"]: q for q in questions}
    tags: List[str] = []
    cal_target = body.base_calories or 2000
    level = "beginner"
    for ans in body.answers:
        q = q_map.get(ans.get("question_id"))
        if not q:
            continue
        if q["type"] == "single":
            idx = ans.get("option_index", 0)
            if 0 <= idx < len(q.get("options", [])):
                opt = q["options"][idx]
                tags.extend(opt.get("tags", []))
                cal_target += opt.get("calorie_modifier", 0)
                if opt.get("level"):
                    level = opt["level"]
    foods = await db.foods.find({}, {"_id": 0}).to_list(2000)
    routines = await db.routines.find({}, {"_id": 0}).to_list(500)
    gym = await db.gyms.find_one({"id": user["gym_id"]}, {"_id": 0}) or {}
    available_equip = gym.get("equipment", [])
    diet = _generate_diet(foods, cal_target, tags)
    routine = _select_routine(routines, level, tags, available_equip)
    plan_doc = {
        "member_id": user["id"],
        "gym_id": user["gym_id"],
        "tags": tags,
        "level": level,
        "calorie_target": cal_target,
        "diet": diet,
        "routine": routine,
        "created_at": now_utc().isoformat(),
        "answers": body.answers,
    }
    await db.member_plans.update_one({"member_id": user["id"]}, {"$set": plan_doc}, upsert=True)
    await db.users.update_one({"id": user["id"]}, {"$set": {"level": level, "level_updated_at": now_utc().isoformat()}})
    return plan_doc

@api.get("/member/plan")
async def member_plan(user=Depends(require_role("member"))):
    plan = await db.member_plans.find_one({"member_id": user["id"]}, {"_id": 0})
    if not plan:
        raise HTTPException(404, "Sin plan. Completa el cuestionario primero.")
    return plan

@api.post("/member/level-up")
async def member_level_up(user=Depends(require_role("member"))):
    progression = {"beginner": "intermediate", "intermediate": "advanced", "advanced": "advanced"}
    cur = user.get("level", "beginner")
    new_level = progression.get(cur, "beginner")
    routines = await db.routines.find({}, {"_id": 0}).to_list(500)
    plan = await db.member_plans.find_one({"member_id": user["id"]}, {"_id": 0}) or {}
    gym = await db.gyms.find_one({"id": user["gym_id"]}, {"_id": 0}) or {}
    new_routine = _select_routine(routines, new_level, plan.get("tags", []), gym.get("equipment", []))
    await db.users.update_one({"id": user["id"]}, {"$set": {"level": new_level, "level_updated_at": now_utc().isoformat()}})
    if plan:
        await db.member_plans.update_one({"member_id": user["id"]}, {"$set": {"level": new_level, "routine": new_routine}})
    return {"level": new_level}

@api.get("/member/plan/pdf")
async def member_plan_pdf(user=Depends(require_role("member"))):
    plan = await db.member_plans.find_one({"member_id": user["id"]}, {"_id": 0})
    if not plan:
        raise HTTPException(404, "Sin plan")
    gym = await db.gyms.find_one({"id": user["gym_id"]}, {"_id": 0}) or {}
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Title"], textColor=colors.HexColor("#FF3B30"),
                                 fontSize=22, spaceAfter=12)
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], textColor=colors.HexColor("#0A0A0A"),
                        fontSize=14, spaceAfter=8)
    body_st = styles["BodyText"]
    story = []
    story.append(Paragraph(f"GymBros — Plan Personalizado", title_style))
    story.append(Paragraph(f"<b>Gimnasio:</b> {gym.get('name', '')}", body_st))
    story.append(Paragraph(f"<b>Miembro:</b> {user['name']} ({user['email']})", body_st))
    story.append(Paragraph(f"<b>Nivel:</b> {plan.get('level', '').title()}", body_st))
    story.append(Paragraph(f"<b>Calorías objetivo:</b> {plan.get('calorie_target', 0)} kcal/día", body_st))
    story.append(Spacer(1, 12))

    story.append(Paragraph("Plan de Dieta", h2))
    diet = plan.get("diet", {})
    for meal, items in (diet.get("meals") or {}).items():
        story.append(Paragraph(f"<b>{meal.title()}</b>", body_st))
        if not items:
            story.append(Paragraph("• (Sin alimentos asignados)", body_st))
        else:
            data = [["Alimento", "Cantidad", "Calorías"]]
            for it in items:
                data.append([it.get("name", ""), f"{it.get('amount', 0)} {it.get('unit', '')}", f"{it.get('calories', 0)} kcal"])
            t = Table(data, hAlign="LEFT", colWidths=[7*cm, 4*cm, 3*cm])
            t.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0A0A0A")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
            ]))
            story.append(t)
        story.append(Spacer(1, 8))

    story.append(PageBreak())
    story.append(Paragraph("Rutina de Ejercicios", h2))
    routine = plan.get("routine") or {}
    if routine:
        story.append(Paragraph(f"<b>Rutina:</b> {routine.get('name', '')}", body_st))
        story.append(Paragraph(f"<b>Nivel:</b> {routine.get('level', '')}", body_st))
        story.append(Spacer(1, 8))
        ex_data = [["Ejercicio", "Series", "Reps", "Equipamiento", "Notas"]]
        for ex in routine.get("exercises", []):
            ex_data.append([ex.get("name", ""), str(ex.get("sets", "")), ex.get("reps", ""),
                            ex.get("equipment", ""), ex.get("notes", "")])
        t = Table(ex_data, hAlign="LEFT", colWidths=[5*cm, 1.5*cm, 2*cm, 4*cm, 4*cm])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#FF3B30")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
        ]))
        story.append(t)
    else:
        story.append(Paragraph("Sin rutina asignada todavía.", body_st))
    doc.build(story)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": "attachment; filename=plan.pdf"})

# Mount router
app.include_router(api)
